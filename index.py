from openpyxl import Workbook ,load_workbook
import sqlite3
con = sqlite3.connect("simple.db")



cur = con.cursor()
'''CREATE TABLE atte( id INT PRIMARY KEY NOT NULL, name TEXT NOT NULL);'''
'''PRAGMA table_info('atte')'''
'''INSERT INTO atte VALUES(123456, 'peter england');'''
'''ALTER TABLE atte ADD batch VARCHAR; '''
'''create table master(subjectcode,batch,name);'''
'''
    threeYear three_Subjects | schema regno 77f677 667f76 date
    create table third_all(id int primary key not null,name text not null)
'''


book = load_workbook("superxls.xlsx") 
sheet = book.active 




yearName = "devTable"


def studentSubCodeAndBatchType():
    course = []
    for i in range(3,sheet.max_column+1):
        subjectcode  =  sheet.cell(row = 1, column = i).value
        student  =  sheet.cell(row = 2, column = i).value
        course.append({"ID":subjectcode,"Type":student})
    return course



# inserting year table 

# cur.execute(f"create table {yearName}(regno,name,batch)")

def storeStudentSqlite():
    batch = 1
    for i in range(2,sheet.max_row+1):
        regno =  sheet.cell(row = i, column = 1).value
        name  =  sheet.cell(row = i, column = 2).value
        if(not regno or  not name):
            batch = 2
            continue
        cur.execute(f"INSERT INTO {yearName} VALUES('{regno}','{name}','{batch}');")

def storeTableMetaData(course):
    for i in course:
        ID = i['ID']
        Type = i['Type']
        cur.execute(f"insert into master values('{ID}','{Type}','{yearName}')")

storeStudentSqlite()

course = studentSubCodeAndBatchType()

storeTableMetaData(course)

def fetchStudent(subject):
    res = cur.execute(f"select * from master where subjectcode='{subject}'").fetchone()
    tableName = res[2]
    batch = res[1]
    students = ""
    if batch=="all":
        students = cur.execute(f"select * from {tableName}").fetchall()
    else:
        students = cur.execute(f"select * from {tableName} where batch='{batch}'").fetchall()
    return students


cur.execute(f"delete from master")
cur.execute(f"delete from {yearName}")
con.commit()
con.close()