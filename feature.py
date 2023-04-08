import requests
from openpyxl import Workbook ,load_workbook
import sqlite3
from json import loads,dumps
con = sqlite3.connect("simple.db")
cur = con.cursor()


base_url = r"http://api.telegram.org/bot6015155943:AAFIy762-g2ZhmWzyUug_xTXkSOmL45aklo"

key = "6015155943:AAFIy762-g2ZhmWzyUug_xTXkSOmL45aklo"


def send_msg(uid,msg):
    parameters = {
      "chat_id" : uid,
      "text" : msg
    }
    resp = requests.get(base_url + "/sendMessage", params=parameters)
    print(resp.text)

def fetchStudent(subject):
    res = cur.execute(f"select * from master where subjectcode='{subject}'").fetchone()
    print("QUERY subject:",res)
    if(res==None):return False
    tableName = res[2]
    batch = res[1]
    students = ""
    if batch=="all":
        students = cur.execute(f"select * from {tableName}").fetchall()
    else:
        students = cur.execute(f"select * from {tableName} where batch='{batch}'").fetchall()
    print(students)
    return students

def fetchStudentData(tabName):
    book = load_workbook(f"static/{tabName}.xlsx") 
    sheet = book.active 
    cur.execute(f"create table {tabName}(regno,name,batch)")
    def studentSubCodeAndBatchType():
        course = []
        for i in range(3,sheet.max_column+1):
            subjectcode  =  sheet.cell(row = 1, column = i).value
            student  =  sheet.cell(row = 2, column = i).value
            course.append({"ID":subjectcode,"Type":student})
        return course

    def storeStudentSqlite():
        batch = 1
        for i in range(2,sheet.max_row+1):
            regno =  sheet.cell(row = i, column = 1).value
            name  =  sheet.cell(row = i, column = 2).value
            if(not regno or  not name):
                batch = 2
                continue
            cur.execute(f"INSERT INTO {tabName} VALUES('{regno}','{name}','{batch}');")

    def storeTableMetaData(course):
        for i in course:
            ID = i['ID']
            Type = i['Type']
            cur.execute(f"insert into master values('{ID}','{Type}','{tabName}')")

    storeStudentSqlite()

    course = studentSubCodeAndBatchType()

    storeTableMetaData(course)




def receiveImage(uid,dict1):
    
    if(dict1["message"].get("document","NULL")!="NULL"):
        FILE_URL = dict1["message"]["document"]
    else:
        FILE_URL = dict1["message"]["photo"][2] 
    print(dumps(dict1,indent=4))
    caption = dict1["message"]["caption"]
    url = requests.get('https://api.telegram.org/bot'+str(key)+"/getFile?file_id="+str(FILE_URL['file_id']))
    path = loads(url.content)["result"]["file_path"]
    img = requests.get('https://api.telegram.org/file/bot'+str(key)+'/'+path)

    with open(f"static/{dict1['message']['caption']}.xlsx","wb") as file:
        file.write(img.content)
    fetchStudentData(caption)

users = [ ]

liveUser = []

askUser = []
currentIndex = 0

students = ["ram","peter","england","john wick","iron man"]

index = 0

def read_msg(offset):
  global index , askUser,currentIndex
  parameters = {
      "offset" : offset
  }

  resp = requests.get(base_url + "/getUpdates", params= parameters)
  data = resp.json()

  print(dumps(data,indent=4))



  for result in data["result"]:
     ID = result["message"]["chat"]["id"]
     text = result["message"].get("text","false").lower()
     
     print(ID,text)
     if ID in users or ID in liveUser:
        if text=="false":
            send_msg(ID, "Processing file...")
            receiveImage(ID, result)
            users.remove(ID)
            send_msg(ID, "processing done")
        elif ID in liveUser:
            if currentIndex < len(askUser):
                send_msg(ID,askUser[currentIndex][1])
                currentIndex+=1
            else:
                currentIndex = 0
                liveUser.remove(ID)
                send_msg(ID,"saved")
            
        else:
            users.remove(ID)
            send_msg(ID,"not valid")
     elif(text=="hi"):
        send_msg(ID, "Hello")
     elif text == "upload":
        send_msg(ID, "send the Excel File and give named caption")
        users.append(ID)
     elif text.split()[0]== "get":
        subjectCode = text.split()[1].upper()
        print("Query: ",subjectCode)
        askUser = fetchStudent(subjectCode)
        if askUser:
            send_msg(ID,f"{len(askUser)} students found")
            send_msg(ID,askUser[currentIndex][1])
            currentIndex+=1
            liveUser.append(ID)
        else:
            send_msg(ID,"students not found")
        
     else:
        send_msg(ID, "Try again")   

     if data["result"]:
       return data["result"][-1]["update_id"] + 1

offset = 0

while True:
  offset = read_msg(offset)
  con.commit()